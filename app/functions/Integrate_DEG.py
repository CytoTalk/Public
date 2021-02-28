from flask import current_app


# print(current_app.config)


def integrate_DEG_cluster_wise(folder_path: str, adj_p: float = 0.05, logFC_cutoff: float = 0.2,
                               fill_zeros: bool = False):
    import openpyxl
    from openpyxl.styles import Color
    import os
    import glob
    print(current_app.config)

    my_red = openpyxl.styles.colors.Color(rgb='ffcccc')
    my_green = openpyxl.styles.colors.Color(rgb='ccffcc')
    my_red_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
    my_green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_green)

    to_exclude = len(folder_path)
    deg_all_files = {}
    all_genes = []

    for filepath in glob.glob(os.path.join(current_app.config['PROJECT_PATH'], folder_path, '*.xlsx')):
        current_file_name = filepath[to_exclude + 1:-5]
        deg_all_files[current_file_name] = {}

    for filepath in glob.glob(os.path.join(current_app.config['PROJECT_PATH'], folder_path, '*.xlsx')):
        current_file_name = filepath[to_exclude + 1:-5]

        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        c = sheet.cell

        for i in range(2, 1000000):
            gene = c(row=i, column=1).value
            if gene is not None:
                gene_adj_p = c(row=i, column=6).value
                gene_logFC = c(row=i, column=3).value
                if gene_adj_p < adj_p:
                    if gene_logFC > logFC_cutoff or gene_logFC == logFC_cutoff:
                        deg_all_files[current_file_name][gene] = []
                        deg_all_files[current_file_name][gene].append(gene_logFC)
                        deg_all_files[current_file_name][gene].append('up')
                        all_genes.append(gene)
                    elif gene_logFC < -logFC_cutoff or gene_logFC == -logFC_cutoff:
                        deg_all_files[current_file_name][gene] = []
                        deg_all_files[current_file_name][gene].append(gene_logFC)
                        deg_all_files[current_file_name][gene].append('down')
                        all_genes.append(gene)
                    else:
                        continue
            else:
                break

    all_genes = set(all_genes)
    all_genes = list(all_genes)

    wb = openpyxl.Workbook()
    sheet = wb.active
    c = sheet.cell

    c(row=1, column=1).value = 'Gene_name'
    col = 2
    for j in deg_all_files:
        c(row=1, column=col).value = j
        col += 1
    r = 2
    for i in all_genes:
        c(row=r, column=1).value = i
        r += 1

    c(row=1, column=len(deg_all_files) + 2).value = 'Upregulated'
    c(row=1, column=len(deg_all_files) + 3).value = 'Downregulated'
    c(row=1, column=len(deg_all_files) + 4).value = 'Bidirectional'

    for i in range(2, len(all_genes) + 2):
        current_gene = c(row=i, column=1).value
        up = 0
        down = 0
        for j in range(2, len(deg_all_files) + 2):
            current_celltype = c(row=1, column=j).value
            if current_gene in deg_all_files[current_celltype]:
                current_logFC = deg_all_files[current_celltype][current_gene][0]
                c(row=i, column=j).value = current_logFC
                if current_logFC > 0:
                    c(row=i, column=j).fill = my_red_fill
                    up += 1
                elif current_logFC < 0:
                    c(row=i, column=j).fill = my_green_fill
                    down += 1
                else:
                    continue
            else:
                if fill_zeros == True:
                    c(row=i, column=j).value = 0
                else:
                    continue

        c(row=i, column=len(deg_all_files) + 2).value = up
        c(row=i, column=len(deg_all_files) + 3).value = down
        if up == 0 or down == 0:
            c(row=i, column=len(deg_all_files) + 4).value = 'No'
        else:
            c(row=i, column=len(deg_all_files) + 4).value = 'Yes'
    # import pdb;
    # pdb.set_trace()
    wb.save(os.path.join(current_app.config['OUTPUT_PATH'], 'Integrated_adj p ' + str(adj_p) + '_LogFC ' + str(
        logFC_cutoff) + '.xlsx'))

# severe_healthy_path = 'assets/Severe_vs_Healthy'
# severe_vs_healthy = integrate_DEG_cluster_wise(severe_healthy_path, adj_p=0.05, logFC_cutoff=0.2, fill_zeros=True)
