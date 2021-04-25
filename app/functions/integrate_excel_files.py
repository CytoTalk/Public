import os
import uuid

from flask import current_app


def integrate_excel(from_file1_path: str, to_file2_path: str, columns_of_interest_from_file1=None,
                    key_from_file1: int = 0,
                    key_to_file2: int = 0,
                    starting_column_file2: int = 0, new_column_name: str = 'Merged', add_annotation: str = "No",
                    add_report: bool = False) -> str:
    """
    :param add_report:
    :param from_file1_path: the path of file that contain information we need to modify_access to the other file
    :param to_file2_path: the path of the file that we will paste information to
    :param columns_of_interest_from_file1: a list of columns numbers to be added from file 1
    :param key_from_file1: the number of the key column in file1 (from_file)
    :param key_to_file2: the number of the key column in file2 (to_file)
    :param starting_column_file2: the number of the column we will start pasting new information into
    :param new_column_name: the name of the new column
    :param add_annotation: if "Gene location" is selected, that would modify_access an extra column after merging the 2 files in a new column

    :return:
    """
    if columns_of_interest_from_file1 is None:
        columns_of_interest_from_file1 = []
    config = current_app.config

    """
    Check User errors
    """
    if len(columns_of_interest_from_file1) < 1:
        return 'You should specify the numbers of columns of interest from file 1 to be added to file 2.'
    for i in columns_of_interest_from_file1:
        if type(i) != int:
            return 'The number of each column of interest should be an integer'
        if i < 0:
            return 'The number of each column of interest should be greater than 0'
    if type(key_to_file2) != int or type(key_from_file1) != int:
        return 'The key column should be the number of the column to be a key to match the 2 files and should be ' \
               'greated than zero '
    if key_to_file2 < 1 or key_from_file1 < 1:
        return 'The number of the key column must be at least 1'

    # load the package that will open the excel files
    import openpyxl

    # open the first file from which we will copy information
    wb1 = openpyxl.load_workbook(from_file1_path, data_only=True)
    sheet1 = wb1.active
    c1 = sheet1.cell

    # open the second file to which we will paste information
    wb2 = openpyxl.load_workbook(to_file2_path, data_only=True)
    sheet2 = wb2.active
    c2 = sheet2.cell

    """
    store the data from file 1 in a dictionary (key is "from_file1_path" key column and value is a list from the columns of interest)
    """
    file1_data = {}
    # get all the data from all rows (i)
    for i in range(2, 1000000):
        key = c1(row=i, column=key_from_file1).value
        if key is None:  # if the ~row is empty (~end of file)
            break
        else:
            file1_data[key] = []
            # Iterate over the columns of interest and append their data in a list called 'value'
            for j in columns_of_interest_from_file1:
                file1_data[key].append(c1(row=i, column=j).value)
    # Now, we have a dictionary "file1_data" to be used to modify_access information to the other file

    """
    Paste the data to file 2
    """

    # First, name the new columns with the name the user entered as an argument
    col = starting_column_file2
    for j in range(0, len(columns_of_interest_from_file1)):
        c2(row=1, column=starting_column_file2 + j).value = new_column_name + '_' + str(j)

    # Check all rows using the key column to match key from file1 with key from file 2
    for i in range(2, 10000000):
        key = c2(row=i, column=key_to_file2).value
        if key is None:
            break
        else:
            if key in file1_data:
                for j in range(0, len(file1_data[key])):
                    c2(row=i, column=starting_column_file2 + j).value = file1_data[key][j]

    """
    If the user is interested to modify_access gene_location
    """

    if add_annotation == "No":
        pass

    if add_annotation == "Gene location":
        # reading from external file (not the user files)
        wb_location = openpyxl.load_workbook(os.path.join(config['PROJECT_PATH'], 'assets', 'locations.xlsx'),
                                             data_only=True)
        sheet_location = wb_location.active
        c_location = sheet_location.cell

        locations = {}
        for i in range(2, 38944):
            locations[c_location(row=i, column=1).value] = c_location(row=i, column=2).value

        c2(row=1, column=starting_column_file2 + len(columns_of_interest_from_file1)).value = "Gene location"
        for i in range(2, 1000000):
            key = c2(row=i, column=key_to_file2).value
            if key is None:
                break
            else:
                if key in locations:
                    c2(row=i, column=starting_column_file2 + len(columns_of_interest_from_file1)).value = locations[key]
    # saving the 2 files
    output_name = str(uuid.uuid4())
    wb2.save(os.path.join(config['OUTPUT_PATH'], output_name + '.xlsx'))
    if add_report:
        with open(os.path.join(config['OUTPUT_PATH'], 'reports', output_name + '.txt'), "w+") as writer:
            writer.write('Analysis Report will be written here')
    return output_name

# from_file = '/home/hp/Desktop/COVID blood samples/from_file1_path.xlsx'
# to_file = '/home/hp/Desktop/COVID blood samples/past_to.xlsx'
#
# integrate_excel(from_file, to_file, columns_of_interest_from_file1=[3, 5], key_from_file1=1, key_to_file2=1,
#                 starting_column_file2=3, new_column_name='Test', add_annotation="Gene location", add_report=True)
